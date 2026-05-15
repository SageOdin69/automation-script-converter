"""
file: ui_element_detector.py

Detects, classifies, and annotates UI elements in HSI screenshots.
Supports both dynamic detection (original) and template‑based detection.
Can output grayscale images of inputs.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import cv2
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False
    warnings.warn("openpyxl not found. Install with: pip install openpyxl. "
                  "Excel export will be skipped.", stacklevel=2)

try:
    from rapidocr_onnxruntime import RapidOCR as _RapidOCR
    _OCR_AVAILABLE = True
except ImportError:
    _OCR_AVAILABLE = False
    warnings.warn("rapidocr-onnxruntime not found. "
                  "Install with: pip install rapidocr-onnxruntime", stacklevel=2)


# =============================================================================
# Colour constants (BGR)
# =============================================================================

TYPE_COLORS: Dict[str, Tuple[int, int, int]] = {
    "BUTTON_ACTIVE":     (0,   200,  80),
    "BUTTON_NORMAL":     (200, 160,  40),
    "BUTTON_ALERT":      (0,   140, 255),
    "TAB_ACTIVE":        (0,   230, 180),
    "TAB_INACTIVE":      (160, 160, 160),
    "SECTION_LABEL":     (255, 200,   0),
    "TITLE":             (255, 255, 255),
    "DISPLAY_FIELD":     (140,  80, 200),
    "AIRCRAFT_SYMBOL":   (80,  200, 220),
    "NAVBAR_ITEM":       (180, 100, 255),
    "PANEL_HEADER":      (100, 210, 255),
    "TEXT_LABEL":        (190, 190, 190),
    "ATTENDANT_ENTRY":   (50,  180, 255),
    "SLIDER_WIDGET":     (200, 200,   0),
    "SCENARIO_BUTTON":   (200, 120, 120),
    "LIGHT_STRIP_ITEM":  (220, 180, 100),
    "MAINTENANCE_ITEM":  (200, 100, 200),
    "LAYER_BUTTON":      (180, 180, 255),
    "ZONE_TAB":          (0,   255, 200),
    "UNKNOWN":           (100, 100, 100),
}

GREEN_LOW  = np.array([40,  80,  80], dtype=np.uint8)
GREEN_HIGH = np.array([90, 255, 255], dtype=np.uint8)
ORANGE_LOW = np.array([10,  80, 180], dtype=np.uint8)
ORANGE_HIGH= np.array([25, 255, 255], dtype=np.uint8)


# =============================================================================
# Data model
# =============================================================================

@dataclass
class UIElement:
    idx:          int
    element_type: str
    label:        str
    x1:           int
    y1:           int
    x2:           int
    y2:           int
    confidence:   float
    crop_path:    str = ""
    notes:        str = ""

    @property
    def width(self) -> int: return self.x2 - self.x1
    @property
    def height(self) -> int: return self.y2 - self.y1
    @property
    def area(self) -> int: return self.width * self.height

    def to_dict(self) -> dict:
        return {
            "idx": self.idx,
            "element_type": self.element_type,
            "label": self.label,
            "x1": self.x1, "y1": self.y1, "x2": self.x2, "y2": self.y2,
            "confidence": self.confidence,
            "notes": self.notes
        }

    @classmethod
    def from_dict(cls, d: dict, idx: int = None):
        return cls(
            idx = idx if idx is not None else d["idx"],
            element_type = d["element_type"],
            label = d["label"],
            x1 = d["x1"], y1 = d["y1"], x2 = d["x2"], y2 = d["y2"],
            confidence = d["confidence"],
            notes = d.get("notes", "")
        )


# =============================================================================
# Classifier (dynamic aircraft detection)
# =============================================================================

class ElementClassifier:
    def __init__(self, img_bgr: np.ndarray, img_h: int, img_w: int):
        self._img = img_bgr
        self._hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
        self._h, self._w = img_h, img_w

    def classify(self, x1: int, y1: int, x2: int, y2: int, text: str = "") -> str:
        cx = (x1 + x2) / 2
        cy = (y1 + y2) / 2
        w, h = x2 - x1, y2 - y1
        txt_lower = text.lower()

        in_topbar      = y2 < 60
        in_title       = y1 < 50 and cx > 150
        in_left_panel  = x2 < 280
        in_entries_panel = (x1 > 350 and y1 > 120 and x2 < 700)
        in_local_panel = x1 > 680 and y1 > 100
        in_bottom_nav  = y1 > self._h * 0.85
        in_aircraft_zone = (200 < cx < self._w - 200 and
                             100 < cy < self._h - 100 and
                             w > 80 and h > 100)

        roi = self._roi(x1, y1, x2, y2)
        is_green   = self._in_range(roi, GREEN_LOW, GREEN_HIGH)
        is_orange  = self._in_range(roi, ORANGE_LOW, ORANGE_HIGH)
        brightness = self._avg_color(roi)[2]

        if in_topbar:
            return "BUTTON_ALERT" if "caution" in txt_lower else "BUTTON_NORMAL"
        if in_title and len(text) > 3:
            return "TITLE"
        if in_aircraft_zone and w > 80 and h > 100 and brightness < 80 and not text:
            return "AIRCRAFT_SYMBOL"
        if "zone" in txt_lower and h < 35:
            return "TAB_ACTIVE" if is_green else "TAB_INACTIVE"
        if in_bottom_nav:
            return "LAYER_BUTTON"
        if "entire scenario" in txt_lower or ("select" in txt_lower and w > 80):
            return "SCENARIO_BUTTON"
        if re.search(r"sect\.\d", txt_lower) or "light strip" in txt_lower:
            return "LIGHT_STRIP_ITEM"
        if "maint" in txt_lower:
            return "MAINTENANCE_ITEM"
        if in_left_panel and h < 30 and not is_green:
            section_hints = ["general", "entire", "light", "maintenance",
                             "day/night", "main on/off", "lights"]
            if any(s in txt_lower for s in section_hints) or (text and len(text) < 20):
                return "SECTION_LABEL"
        if in_entries_panel and w > 80 and h > 25:
            return "ATTENDANT_ENTRY"
        if in_local_panel and ((w < 60 and h > 20) or "brt" in txt_lower or "dim" in txt_lower):
            return "SLIDER_WIDGET"
        if w > 30 and h > 20:
            if is_green:   return "BUTTON_ACTIVE"
            if is_orange:  return "BUTTON_ALERT"
            if brightness > 40: return "BUTTON_NORMAL"
        if (in_entries_panel or in_local_panel) and h < 25:
            return "PANEL_HEADER"
        if h < 22:
            return "TEXT_LABEL"
        return "UNKNOWN"

    def _roi(self, x1, y1, x2, y2) -> np.ndarray:
        pad = 4
        y1c = max(0, y1 + pad)
        y2c = min(self._h, y2 - pad)
        x1c = max(0, x1 + pad)
        x2c = min(self._w, x2 - pad)
        if y2c <= y1c or x2c <= x1c:
            return self._hsv[y1:y2, x1:x2]
        return self._hsv[y1c:y2c, x1c:x2c]

    @staticmethod
    def _avg_color(roi: np.ndarray) -> np.ndarray:
        if roi.size == 0:
            return np.zeros(3)
        return roi.reshape(-1, 3).mean(axis=0)

    @staticmethod
    def _in_range(roi: np.ndarray, lo, hi, threshold=0.12) -> bool:
        if roi.size == 0:
            return False
        mask = cv2.inRange(roi, lo, hi)
        return mask.mean() / 255.0 > threshold


# =============================================================================
# OCR and Contour builders
# =============================================================================

class OCRElementBuilder:
    def __init__(self):
        self._ocr = _RapidOCR() if _OCR_AVAILABLE else None

    def extract(self, img: np.ndarray) -> List[Dict]:
        if not self._ocr:
            return []
        result, _ = self._ocr(img)
        if not result:
            return []
        hits = []
        for item in result:
            box, text, conf = item[0], item[1], float(item[2])
            xs = [int(p[0]) for p in box]
            ys = [int(p[1]) for p in box]
            hits.append({"text": text, "x1": min(xs), "y1": min(ys),
                         "x2": max(xs), "y2": max(ys), "conf": conf})
        return hits


class ContourElementBuilder:
    MIN_AREA = 600
    MAX_AREA_RATIO = 0.5
    MIN_ASPECT = 0.1
    MAX_ASPECT = 20.0

    def extract(self, img: np.ndarray) -> List[Dict]:
        h, w = img.shape[:2]
        max_area = h * w * self.MAX_AREA_RATIO
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blur = cv2.bilateralFilter(gray, 9, 75, 75)
        edges = cv2.Canny(blur, 30, 90)
        kernel = np.ones((3, 3), np.uint8)
        edges = cv2.dilate(edges, kernel, iterations=1)
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        hits = []
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < self.MIN_AREA or area > max_area:
                continue
            x, y, cw, ch = cv2.boundingRect(cnt)
            aspect = cw / max(ch, 1)
            if not (self.MIN_ASPECT <= aspect <= self.MAX_ASPECT):
                continue
            hits.append({"text": "", "x1": x, "y1": y, "x2": x + cw, "y2": y + ch, "conf": 0.70})
        return hits


# =============================================================================
# NMS
# =============================================================================

def nms(elements: List[UIElement], iou_threshold: float = 0.45) -> List[UIElement]:
    if not elements:
        return []
    boxes = [(e.x1, e.y1, e.x2, e.y2) for e in elements]
    scores = [e.confidence for e in elements]
    order = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)
    keep = []
    while order:
        i = order.pop(0)
        keep.append(i)
        order = [j for j in order if _iou(boxes[i], boxes[j]) < iou_threshold]
    return [elements[i] for i in keep]

def _iou(a, b):
    ix1 = max(a[0], b[0]); iy1 = max(a[1], b[1])
    ix2 = min(a[2], b[2]); iy2 = min(a[3], b[3])
    inter = max(0, ix2 - ix1) * max(0, iy2 - iy1)
    if inter == 0:
        return 0.0
    ua = (a[2]-a[0])*(a[3]-a[1]) + (b[2]-b[0])*(b[3]-b[1]) - inter
    return inter / max(ua, 1)


# =============================================================================
# Main detector (dynamic + template support)
# =============================================================================

class UIElementDetector:
    BUTTON_PAD = 12
    LABEL_PAD = 4

    def __init__(self):
        self._ocr_builder = OCRElementBuilder()
        self._contour_builder = ContourElementBuilder()

    # -------------------------------------------------------------------------
    # Dynamic detection (original)
    # -------------------------------------------------------------------------
    def detect_dynamic(self, image_path: str, out_dir: str = "./output",
                       save_crops: bool = True, save_annotated: bool = True,
                       save_xlsx: bool = True, save_csv: bool = True,
                       output_grayscale: bool = False,
                       extract_template: bool = False) -> List[UIElement]:
        img = cv2.imread(image_path)
        if img is None:
            raise FileNotFoundError(f"Cannot read: {image_path}")
        h, w = img.shape[:2]

        out_path = Path(out_dir)
        out_path.mkdir(parents=True, exist_ok=True)

        # Optional grayscale output
        if output_grayscale:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            gray_path = out_path / f"grayscale_{Path(image_path).name}"
            cv2.imwrite(str(gray_path), gray)
            print(f"      Grayscale saved → {gray_path}")

        print(f"\n[1/5] Running OCR on {Path(image_path).name} ({w}×{h}) …")
        raw_ocr = self._ocr_builder.extract(img)
        print(f"      {len(raw_ocr)} raw OCR hits")

        print("[2/5] Running contour detection …")
        contours = self._contour_builder.extract(img)
        print(f"      {len(contours)} contour candidates")

        print("[3/5] Classifying elements …")
        clf = ElementClassifier(img, h, w)
        elements = self._build_elements(raw_ocr, contours, clf, img, w, h)
        print(f"      {len(elements)} elements after NMS")

        if save_crops:
            print("[4/5] Saving crops …")
            self._save_crops(img, elements, out_path)

        if save_annotated:
            print("[5a] Saving annotated image …")
            ann_path = out_path / f"annotated_{Path(image_path).name}"
            self._annotate(img.copy(), elements, str(ann_path))

            print("[5b] Saving text-annotated image …")
            txt_ann_path = out_path / f"text_annotated_{Path(image_path).name}"
            self._annotate_text(img.copy(), raw_ocr, str(txt_ann_path))

        if save_xlsx and _XLSX_AVAILABLE:
            print("[5c] Saving Excel ROI datasheet …")
            self._save_xlsx(elements, out_path / "roi_datasheet.xlsx", Path(image_path).name)

        if save_csv:
            print("[5d] Saving CSV ROI datasheet …")
            self._save_csv(elements, out_path / "roi_datasheet.csv")

        if raw_ocr:
            print("[5e] Saving text datasheet …")
            self._save_text_datasheet(raw_ocr, out_path)

        if extract_template:
            template_path = out_path / "template.json"
            self._save_template(elements, template_path)
            print(f"[5f] Template saved → {template_path}")

        # Summary
        from collections import Counter
        counts = Counter(e.element_type for e in elements)
        print("\n  ── Detection summary ──────────────────────")
        for t, c in sorted(counts.items()):
            print(f"     {t:<22} {c:>3} element(s)")
        print(f"  ── Total: {len(elements)} ─────────────────────────\n")
        return elements

    # -------------------------------------------------------------------------
    # Template‑based detection (crop ROIs + OCR only)
    # -------------------------------------------------------------------------
    def detect_from_template(self, image_path: str, template_file: str,
                             out_dir: str = "./output",
                             save_crops: bool = True,
                             output_grayscale: bool = False) -> List[UIElement]:
        img = cv2.imread(image_path)
        if img is None:
            raise FileNotFoundError(f"Cannot read: {image_path}")
        h, w = img.shape[:2]

        out_path = Path(out_dir)
        out_path.mkdir(parents=True, exist_ok=True)

        # Optional grayscale output
        if output_grayscale:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            gray_path = out_path / f"grayscale_{Path(image_path).name}"
            cv2.imwrite(str(gray_path), gray)
            print(f"      Grayscale saved → {gray_path}")

        # Load template
        with open(template_file, "r") as f:
            template_data = json.load(f)
        elements = [UIElement.from_dict(d, idx=i+1) for i, d in enumerate(template_data)]

        print(f"\n[Template] Loaded {len(elements)} ROIs from {template_file}")

        # Run OCR on each ROI
        ocr = _RapidOCR() if _OCR_AVAILABLE else None
        ocr_hits = []   # for text annotation

        for e in elements:
            x1, y1, x2, y2 = e.x1, e.y1, e.x2, e.y2
            crop = img[y1:y2, x1:x2]
            # Read text inside crop
            if ocr:
                result, _ = ocr(crop)
                if result:
                    texts = [item[1] for item in result]
                    e.label = " ".join(texts)[:60]
                    e.notes = "template‑based (OCR only)"
                    # Store for text annotation
                    for item in result:
                        box, text, conf = item[0], item[1], item[2]
                        xs = [int(p[0]) for p in box]
                        ys = [int(p[1]) for p in box]
                        ocr_hits.append({
                            "text": text,
                            "x1": x1 + min(xs), "y1": y1 + min(ys),
                            "x2": x1 + max(xs), "y2": y1 + max(ys),
                            "conf": float(conf)
                        })
                else:
                    e.label = ""
            else:
                e.label = ""
            e.confidence = 1.0  # trust template

        if save_crops:
            print("[Template] Saving crops …")
            self._save_crops(img, elements, out_path)

        # Annotated image (boundaries) – optional but we'll do it
        print("[Template] Saving annotated image …")
        ann_path = out_path / f"annotated_{Path(image_path).name}"
        self._annotate(img.copy(), elements, str(ann_path))

        if ocr_hits:
            print("[Template] Saving text-annotated image …")
            txt_ann_path = out_path / f"text_annotated_{Path(image_path).name}"
            self._annotate_text(img.copy(), ocr_hits, str(txt_ann_path))

        # Save CSV datasheet for this run
        self._save_csv(elements, out_path / "roi_datasheet.csv")

        # Summary
        from collections import Counter
        counts = Counter(e.element_type for e in elements)
        print("\n  ── Template detection summary ─────────────────")
        for t, c in sorted(counts.items()):
            print(f"     {t:<22} {c:>3} element(s)")
        print(f"  ── Total: {len(elements)} ─────────────────────────\n")
        return elements

    # -------------------------------------------------------------------------
    # Internal builders (unchanged)
    # -------------------------------------------------------------------------
    def _build_elements(self, ocr_hits, contours, clf, img, img_w, img_h):
        raw_elements = []
        idx = 0

        for hit in ocr_hits:
            x1, y1, x2, y2 = hit["x1"], hit["y1"], hit["x2"], hit["y2"]
            text = hit["text"].strip()
            pad = self.LABEL_PAD if len(text) > 20 else self.BUTTON_PAD
            x1e, y1e = max(0, x1 - pad), max(0, y1 - pad)
            x2e, y2e = min(img_w, x2 + pad), min(img_h, y2 + pad)
            etype = clf.classify(x1e, y1e, x2e, y2e, text)
            raw_elements.append(UIElement(
                idx=idx, element_type=etype, label=self._clean_label(text),
                x1=x1e, y1=y1e, x2=x2e, y2=y2e, confidence=hit["conf"]
            ))
            idx += 1

        for c in contours:
            x1, y1, x2, y2 = c["x1"], c["y1"], c["x2"], c["y2"]
            if any(_iou((x1,y1,x2,y2), (e.x1,e.y1,e.x2,e.y2)) > 0.35 for e in raw_elements):
                continue
            etype = clf.classify(x1, y1, x2, y2, "")
            if etype == "UNKNOWN":
                continue
            raw_elements.append(UIElement(
                idx=idx, element_type=etype, label=f"{etype.lower()}_region",
                x1=x1, y1=y1, x2=x2, y2=y2, confidence=c["conf"], notes="contour-detected"
            ))
            idx += 1

        final = nms(raw_elements, iou_threshold=0.50)
        for i, e in enumerate(final):
            e.idx = i + 1
        return final

    # -------------------------------------------------------------------------
    # Template saving
    # -------------------------------------------------------------------------
    def _save_template(self, elements: List[UIElement], path: Path):
        data = [e.to_dict() for e in elements]
        with open(path, "w") as f:
            json.dump(data, f, indent=2)

    # -------------------------------------------------------------------------
    # Annotation and export (unchanged from original, but included for completeness)
    # -------------------------------------------------------------------------
    def _annotate(self, img: np.ndarray, elements: List[UIElement], out_path: str) -> None:
        overlay = img.copy()
        for e in elements:
            color = TYPE_COLORS.get(e.element_type, (100,100,100))
            cv2.rectangle(overlay, (e.x1, e.y1), (e.x2, e.y2), color, -1)
        cv2.addWeighted(overlay, 0.18, img, 0.82, 0, img)
        for e in elements:
            color = TYPE_COLORS.get(e.element_type, (100,100,100))
            cv2.rectangle(img, (e.x1, e.y1), (e.x2, e.y2), color, 2)
            tag_x, tag_y = e.x1, max(12, e.y1 - 4)
            tag_text = str(e.idx)
            (tw, th), _ = cv2.getTextSize(tag_text, cv2.FONT_HERSHEY_SIMPLEX, 0.42, 1)
            cv2.rectangle(img, (tag_x, tag_y - th - 3), (tag_x + tw + 4, tag_y + 2), color, -1)
            cv2.putText(img, tag_text, (tag_x + 2, tag_y - 1),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0,0,0), 1, cv2.LINE_AA)
        cv2.imwrite(out_path, img)

    def _annotate_text(self, img: np.ndarray, ocr_hits: List[Dict], out_path: str) -> None:
        overlay = img.copy()
        for hit in ocr_hits:
            x1, y1, x2, y2 = hit["x1"], hit["y1"], hit["x2"], hit["y2"]
            cv2.rectangle(overlay, (x1, y1), (x2, y2), (255,255,0), -1)
        cv2.addWeighted(overlay, 0.15, img, 0.85, 0, img)
        for hit in ocr_hits:
            x1, y1, x2, y2 = hit["x1"], hit["y1"], hit["x2"], hit["y2"]
            cv2.putText(img, hit["text"], (x1, max(y1-4,10)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0,0,0), 2, cv2.LINE_AA)
        cv2.imwrite(out_path, img)

    def _save_crops(self, img: np.ndarray, elements: List[UIElement], out_path: Path) -> None:
        crops_dir = out_path / "crops"
        crops_dir.mkdir(exist_ok=True)
        for e in elements:
            type_dir = crops_dir / e.element_type
            type_dir.mkdir(exist_ok=True)
            x1, y1 = max(0, e.x1), max(0, e.y1)
            x2, y2 = min(img.shape[1], e.x2), min(img.shape[0], e.y2)
            if x2 <= x1 or y2 <= y1:
                continue
            crop = img[y1:y2, x1:x2]
            safe = re.sub(r"[^\w]", "_", e.label)[:40]
            fname = f"{e.idx:03d}_{safe}.png"
            fpath = type_dir / fname
            cv2.imwrite(str(fpath), crop)
            e.crop_path = str(fpath)
        print(f"      {sum(1 for e in elements if e.crop_path)} crops saved under {crops_dir}/")

    def _save_text_datasheet(self, ocr_hits: List[Dict], out_path: Path) -> None:
        csv_path = out_path / "text_datasheet.csv"
        headers = ["index","text","x1","y1","x2","y2","confidence"]
        with open(str(csv_path), "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for i, hit in enumerate(ocr_hits, 1):
                writer.writerow({"index": i, "text": hit["text"], "x1": hit["x1"], "y1": hit["y1"],
                                 "x2": hit["x2"], "y2": hit["y2"], "confidence": round(hit["conf"],3)})
        print(f"      Saved → {csv_path}")
        if _XLSX_AVAILABLE:
            xlsx_path = out_path / "text_datasheet.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "OCR Text Detections"
            ws.append(headers)
            for i, hit in enumerate(ocr_hits, 1):
                ws.append([i, hit["text"], hit["x1"], hit["y1"], hit["x2"], hit["y2"], round(hit["conf"],3)])
            header_fill = PatternFill("solid", fgColor="1A2B4A")
            header_font = Font(bold=True, color="FFFFFF")
            for col in range(1,8):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            wb.save(str(xlsx_path))
            print(f"      Saved → {xlsx_path}")

    def _save_xlsx(self, elements: List[UIElement], out_path: Path, source_image: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ROI Datasheet"
        headers = ["Index","Type","Label","X1 (px)","Y1 (px)","X2 (px)","Y2 (px)",
                   "Width (px)","Height (px)","Area (px²)","Confidence","Notes","Crop Path"]
        header_fill = PatternFill("solid", fgColor="1A2B4A")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="444444")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, header_align, border
        ws.row_dimensions[1].height = 28

        type_fills = {
            "BUTTON_ACTIVE":"C8F5D5", "BUTTON_NORMAL":"D6E8FF", "BUTTON_ALERT":"FFE0B2",
            "TAB_ACTIVE":"B9F6CA", "TAB_INACTIVE":"EEEEEE", "SECTION_LABEL":"FFF9C4",
            "TITLE":"F3E5F5", "DISPLAY_FIELD":"EDE7F6", "AIRCRAFT_SYMBOL":"E0F7FA",
            "NAVBAR_ITEM":"FCE4EC", "PANEL_HEADER":"E1F5FE", "TEXT_LABEL":"F5F5F5",
            "ATTENDANT_ENTRY":"FFF3E0", "SLIDER_WIDGET":"E8F5E9", "SCENARIO_BUTTON":"FCE4EC",
            "LIGHT_STRIP_ITEM":"FFF0E0", "MAINTENANCE_ITEM":"E8E0F0", "LAYER_BUTTON":"D9E8FF",
            "ZONE_TAB":"C8F7E5", "UNKNOWN":"FAFAFA",
        }
        for row_idx, e in enumerate(elements, 2):
            fill_hex = type_fills.get(e.element_type, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_hex)
            row_data = [e.idx, e.element_type, e.label, e.x1, e.y1, e.x2, e.y2,
                        e.width, e.height, e.area, round(e.confidence,3), e.notes, e.crop_path]
            ws.append(row_data)
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill, cell.border = row_fill, border
                cell.alignment = Alignment(horizontal="center" if col not in (3,13) else "left", vertical="center")
                if col == 11: cell.number_format = "0.000"
                if col == 10: cell.number_format = "#,##0"
        col_widths = [7,20,30,9,9,9,9,10,11,12,12,25,50]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"

        ls = wb.create_sheet("Legend")
        ls.append(["Type","Description","Color code"])
        desc = {
            "BUTTON_ACTIVE":"Pressable button – active (green fill)",
            "BUTTON_NORMAL":"Pressable button – default",
            "BUTTON_ALERT":"Alert button (orange fill)",
            "TAB_ACTIVE":"Selected zone tab",
            "TAB_INACTIVE":"Unselected zone tab",
            "SECTION_LABEL":"Section heading (left panel)",
            "TITLE":"Page title",
            "DISPLAY_FIELD":"Read‑only value field",
            "AIRCRAFT_SYMBOL":"Aircraft silhouette (graphical)",
            "NAVBAR_ITEM":"Bottom navigation bar item",
            "PANEL_HEADER":"Panel heading",
            "TEXT_LABEL":"Small info label",
            "ATTENDANT_ENTRY":"Attendant row in Entries panel",
            "SLIDER_WIDGET":"Brightness/dim slider or button",
            "SCENARIO_BUTTON":"'Entire Scenario' select button",
            "LIGHT_STRIP_ITEM":"Light strip control row",
            "MAINTENANCE_ITEM":"Maintenance entry",
            "LAYER_BUTTON":"Bottom layer navigation button",
            "ZONE_TAB":"Zone selector tab",
            "UNKNOWN":"Unclassified",
        }
        for t, d in desc.items():
            ls.append([t, d, type_fills.get(t,"FFFFFF")])
            row = ls.max_row
            ls.cell(row=row, column=1).fill = PatternFill("solid", fgColor=type_fills.get(t,"FFFFFF"))
        ls.column_dimensions["A"].width = 22
        ls.column_dimensions["B"].width = 45

        ms = wb.create_sheet("Meta")
        ms.append(["Source image", source_image])
        ms.append(["Total elements", len(elements)])
        ms.append(["Generated by", "UIElementDetector (enhanced)"])
        ms.column_dimensions["A"].width = 18
        wb.save(str(out_path))
        print(f"      Saved → {out_path}")

    def _save_csv(self, elements: List[UIElement], out_path: Path) -> None:
        headers = ["index","type","label","x1","y1","x2","y2","width","height","area","confidence","notes","crop_path"]
        with open(str(out_path), "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for e in elements:
                w.writerow({"index":e.idx,"type":e.element_type,"label":e.label,"x1":e.x1,"y1":e.y1,
                            "x2":e.x2,"y2":e.y2,"width":e.width,"height":e.height,"area":e.area,
                            "confidence":round(e.confidence,3),"notes":e.notes,"crop_path":e.crop_path})
        print(f"      Saved → {out_path}")

    @staticmethod
    def _clean_label(text: str) -> str:
        return re.sub(r"\s+", " ", text.strip())[:60]


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Detect UI elements in HSI screenshots (dynamic or template‑based).")
    ap.add_argument("image", help="Path to the input image")
    ap.add_argument("--out-dir", default="./output", help="Output directory")
    ap.add_argument("--no-crops", action="store_true", help="Disable saving individual crops")
    ap.add_argument("--no-annotated", action="store_true", help="Disable saving annotated images")
    ap.add_argument("--no-xlsx", action="store_true", help="Disable Excel export")
    ap.add_argument("--no-csv", action="store_true", help="Disable CSV export")
    ap.add_argument("--output-grayscale", action="store_true", help="Save grayscale version of the input image")
    ap.add_argument("--extract-template", action="store_true", help="After dynamic detection, save ROIs as template.json")
    ap.add_argument("--use-template", type=str, metavar="TEMPLATE.json", help="Use a saved template file instead of dynamic detection")
    args = ap.parse_args()

    detector = UIElementDetector()

    if args.use_template:
        # Template‑based mode
        detector.detect_from_template(
            image_path=args.image,
            template_file=args.use_template,
            out_dir=args.out_dir,
            save_crops=not args.no_crops,
            output_grayscale=args.output_grayscale
        )
    else:
        # Dynamic detection (original behaviour, plus new flags)
        detector.detect_dynamic(
            image_path=args.image,
            out_dir=args.out_dir,
            save_crops=not args.no_crops,
            save_annotated=not args.no_annotated,
            save_xlsx=not args.no_xlsx,
            save_csv=not args.no_csv,
            output_grayscale=args.output_grayscale,
            extract_template=args.extract_template
        )
