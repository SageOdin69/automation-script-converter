"""
file: ui_element_detector.py

Detects, classifies, and annotates every individual UI element in an
HSI reference screenshot. **No multi-line text grouping is applied:**
each OCR text region becomes a separate element.

What it does
------------
1.  OCR pass        — finds every text region with its bounding box
2.  Contour pass    — finds every bordered rectangle (buttons, tabs, panels)
3.  Build elements  — each raw OCR hit becomes one UI element (no merging)
4.  Classify        — assigns an element type using color + shape + position rules
5.  Annotate        — draws colored bounding boxes with type labels on a copy
                      (saves as annotated_<image>)
6.  Text annotation — draws detected text strings on a separate copy
                      (saves as text_annotated_<image>)
7.  Export crops    — saves each element as an individual PNG
8.  Export ROI data — writes a full datasheet to both CSV and Excel (.xlsx)
9.  Export text data— writes a separate datasheet of all OCR text detections

Element types detected
----------------------
BUTTON_ACTIVE     green-fill interactive button (selected state)
BUTTON_NORMAL     dark-fill interactive button
BUTTON_ALERT      orange-fill alert button
TAB_ACTIVE        active tab (green underline)
TAB_INACTIVE      inactive tab
SECTION_LABEL     section heading text (no border)
TITLE             page title text
DISPLAY_FIELD     read-only value field (e.g. "Entire 1.8")
AIRCRAFT_SYMBOL   the aircraft silhouette graphic
NAVBAR_ITEM       bottom navigation bar item
PANEL_HEADER      panel heading inside a bordered panel
TEXT_LABEL        small informational text label
ATTENDANT_ENTRY   attendant row inside the entries panel
SLIDER_WIDGET     the brightness/dim slider widget

Usage
-----
python ui_element_detector.py <image_path> [--out-dir ./output]

or from Python:
    from ui_element_detector import UIElementDetector
    det = UIElementDetector()
    det.detect("3.png", out_dir="./output")
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import cv2
import numpy as np
from PIL import Image

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False
    warnings.warn("openpyxl not found. Install with: pip install openpyxl. "
                  "Excel export will be skipped.", stacklevel=2)

try:
    from rapidocr_onnxruntime import RapidOCR as _RapidOCR
    _OCR_AVAILABLE = True
except ImportError:
    _OCR_AVAILABLE = False
    warnings.warn("rapidocr-onnxruntime not found. "
                  "Install with: pip install rapidocr-onnxruntime", stacklevel=2)


# =============================================================================
# Colour constants (BGR — OpenCV native)
# =============================================================================

# Annotation box colours per element type
TYPE_COLORS: Dict[str, Tuple[int, int, int]] = {
    "BUTTON_ACTIVE":    (0,   200,  80),   # bright green
    "BUTTON_NORMAL":    (200, 160,  40),   # steel blue
    "BUTTON_ALERT":     (0,   140, 255),   # orange
    "TAB_ACTIVE":       (0,   230, 180),   # lime
    "TAB_INACTIVE":     (160, 160, 160),   # grey
    "SECTION_LABEL":    (255, 200,   0),   # cyan
    "TITLE":            (255, 255, 255),   # white
    "DISPLAY_FIELD":    (140,  80, 200),   # purple
    "AIRCRAFT_SYMBOL":  (80,  200, 220),   # teal
    "NAVBAR_ITEM":      (180, 100, 255),   # violet
    "PANEL_HEADER":     (100, 210, 255),   # sky
    "TEXT_LABEL":       (190, 190, 190),   # light grey
    "ATTENDANT_ENTRY":  (50,  180, 255),   # amber
    "SLIDER_WIDGET":    (200, 200,   0),   # aqua
    "UNKNOWN":          (100, 100, 100),   # dark grey
}

# Background color HSV ranges for classification
# HSI screens use a consistent dark-slate theme
GREEN_LOW  = np.array([40,  80,  80],  dtype=np.uint8)
GREEN_HIGH = np.array([90, 255, 255],  dtype=np.uint8)
ORANGE_LOW = np.array([10,  80, 180],  dtype=np.uint8)
ORANGE_HIGH= np.array([25, 255, 255],  dtype=np.uint8)


# =============================================================================
# Data model
# =============================================================================

@dataclass
class UIElement:
    idx:          int
    element_type: str
    label:        str           # best text found inside (or descriptive name)
    x1:           int
    y1:           int
    x2:           int
    y2:           int
    confidence:   float
    crop_path:    str = ""
    notes:        str = ""

    @property
    def roi(self) -> Tuple[int, int, int, int]:
        return (self.x1, self.y1, self.x2, self.y2)

    @property
    def width(self) -> int:
        return self.x2 - self.x1

    @property
    def height(self) -> int:
        return self.y2 - self.y1

    @property
    def area(self) -> int:
        return self.width * self.height


# =============================================================================
# Classifier
# =============================================================================

class ElementClassifier:
    """
    Assigns an element type using:
      - Average HSV color of the region interior
      - Aspect ratio
      - Position in the image (top-bar, bottom-bar, left-panel, main-area)
      - Text content clues
    """

    def __init__(self, img_bgr: np.ndarray, img_h: int, img_w: int):
        self._img  = img_bgr
        self._hsv  = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
        self._h    = img_h
        self._w    = img_w

    def classify(self, x1: int, y1: int, x2: int, y2: int,
                 text: str = "") -> str:
        cx = (x1 + x2) / 2
        cy = (y1 + y2) / 2
        w  = x2 - x1
        h  = y2 - y1

        # ── Position zones ─────────────────────────────────────────────
        in_topbar   = y2 < 60                        # Screen Off / Caution row
        in_title    = y1 < 50 and cx > 200           # page title area
        in_tabs     = 45 < y1 < 90 and cx > 200      # Entries / Zones / Rooms
        in_navbar   = y1 > self._h * 0.90            # bottom nav bar
        in_lpanel   = x2 < 200                       # left control panel
        in_aircraft = (200 < cx < 650 and
                       100 < cy < 700 and
                       w > 60 and h > 100)           # center canvas area
        in_entries  = 380 < x1 < 650 and 130 < y1   # entries panel
        in_local    = x1 > 650 and y1 > 120          # local adjustment

        roi = self._roi(x1, y1, x2, y2)
        avg_color  = self._avg_color(roi)
        is_green   = self._in_range(roi, GREEN_LOW,  GREEN_HIGH)
        is_orange  = self._in_range(roi, ORANGE_LOW, ORANGE_HIGH)
        brightness = int(avg_color[2])               # V channel

        txt_lower = text.lower()

        # ── Classification rules (order matters) ───────────────────────

        if in_title and len(text) > 5:
            return "TITLE"

        if in_topbar:
            if is_orange:
                return "BUTTON_ALERT"
            return "BUTTON_NORMAL"

        if in_tabs:
            return "TAB_ACTIVE" if is_green else "TAB_INACTIVE"

        if in_navbar:
            return "NAVBAR_ITEM"

        # Aircraft symbol: large dark region in the center with no strong text
        if (in_aircraft and w > 80 and h > 200
                and brightness < 80 and not text):
            return "AIRCRAFT_SYMBOL"

        # Section labels (text only, very short height, no button fill)
        if (in_lpanel and h < 25 and not is_green
                and brightness < 90):
            section_hints = ["general", "entire", "light", "maintenance",
                             "aircraft", "strip", "adjustment"]
            if any(s in txt_lower for s in section_hints) or (not text):
                return "SECTION_LABEL"

        # Display field ("Entire 1.8" box)
        if (in_lpanel and w > 60 and h > 20
                and brightness < 80 and not is_green
                and any(c.isdigit() for c in text)):
            return "DISPLAY_FIELD"

        # Slider widget (tall and narrow in local adjustment panel)
        if in_local and w < 50 and h > 80:
            return "SLIDER_WIDGET"

        # Attendant entries
        if in_entries and w > 100 and h > 30:
            return "ATTENDANT_ENTRY" if not is_green else "BUTTON_ACTIVE"

        # Buttons
        if w > 30 and h > 25:
            if is_green:
                return "BUTTON_ACTIVE"
            if is_orange:
                return "BUTTON_ALERT"
            if brightness > 40:
                return "BUTTON_NORMAL"

        # Panel header (inside bordered panel, first text row)
        if in_entries and h < 25:
            return "PANEL_HEADER"

        if in_local and h < 25:
            return "PANEL_HEADER"

        # Small floating text
        if h < 22:
            return "TEXT_LABEL"

        return "UNKNOWN"

    def _roi(self, x1, y1, x2, y2) -> np.ndarray:
        pad = 4
        y1c = max(0, y1 + pad)
        y2c = min(self._h, y2 - pad)
        x1c = max(0, x1 + pad)
        x2c = min(self._w, x2 - pad)
        if y2c <= y1c or x2c <= x1c:
            return self._hsv[max(0,y1):min(self._h,y2),
                             max(0,x1):min(self._w,x2)]
        return self._hsv[y1c:y2c, x1c:x2c]

    @staticmethod
    def _avg_color(roi: np.ndarray) -> np.ndarray:
        if roi.size == 0:
            return np.zeros(3)
        return roi.reshape(-1, 3).mean(axis=0)

    @staticmethod
    def _in_range(roi: np.ndarray, lo: np.ndarray,
                  hi: np.ndarray, threshold: float = 0.12) -> bool:
        if roi.size == 0:
            return False
        mask = cv2.inRange(roi, lo, hi)
        return mask.mean() / 255.0 > threshold


# =============================================================================
# OCR-based element builder (kept for raw extraction only; merging disabled)
# =============================================================================

class OCRElementBuilder:
    """
    Extracts text regions using RapidOCR.
    Merging is no longer used — every raw hit becomes an individual element.
    """

    def __init__(self):
        self._ocr = _RapidOCR() if _OCR_AVAILABLE else None

    def extract(self, img: np.ndarray) -> List[Dict]:
        """Return list of raw dicts: {text, x1,y1,x2,y2, conf}."""
        if not self._ocr:
            return []
        result, _ = self._ocr(img)
        if not result:
            return []
        hits = []
        for item in result:
            box, text, conf = item[0], item[1], float(item[2])
            xs = [int(p[0]) for p in box]
            ys = [int(p[1]) for p in box]
            hits.append({"text": text, "x1": min(xs), "y1": min(ys),
                         "x2": max(xs), "y2": max(ys), "conf": conf})
        return hits

    # The merge method is kept for potential reuse but is no longer called.
    def merge(self, hits: List[Dict]) -> List[Dict]:
        """Not used in the current pipeline."""
        return hits


# =============================================================================
# Contour-based element builder
# =============================================================================

class ContourElementBuilder:
    """
    Uses OpenCV edge detection and contour finding to locate
    bordered rectangles (buttons, tabs, panels) that may have been
    missed by OCR (e.g., empty buttons or the aircraft symbol area).
    """

    MIN_AREA  = 600    # ignore tiny noise contours
    MAX_AREA_RATIO = 0.5  # ignore full-screen-sized contours
    MIN_ASPECT = 0.1
    MAX_ASPECT = 20.0

    def extract(self, img: np.ndarray) -> List[Dict]:
        h, w = img.shape[:2]
        max_area = h * w * self.MAX_AREA_RATIO

        gray  = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # Bilateral filter to preserve edges while reducing noise
        blur  = cv2.bilateralFilter(gray, 9, 75, 75)
        edges = cv2.Canny(blur, 30, 90)
        # Dilate to connect broken edges
        kernel = np.ones((3, 3), np.uint8)
        edges  = cv2.dilate(edges, kernel, iterations=1)

        contours, _ = cv2.findContours(
            edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        hits = []
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < self.MIN_AREA or area > max_area:
                continue
            x, y, cw, ch = cv2.boundingRect(cnt)
            aspect = cw / max(ch, 1)
            if not (self.MIN_ASPECT <= aspect <= self.MAX_ASPECT):
                continue
            hits.append({"text": "", "x1": x, "y1": y,
                         "x2": x + cw, "y2": y + ch, "conf": 0.70})
        return hits


# =============================================================================
# NMS — remove heavily overlapping boxes
# =============================================================================

def nms(elements: List[UIElement],
        iou_threshold: float = 0.45) -> List[UIElement]:
    """Non-maximum suppression by IoU — keep higher-confidence box."""
    if not elements:
        return []
    boxes = [(e.x1, e.y1, e.x2, e.y2) for e in elements]
    scores = [e.confidence for e in elements]
    order  = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)
    keep   = []

    while order:
        i = order.pop(0)
        keep.append(i)
        order = [j for j in order if _iou(boxes[i], boxes[j]) < iou_threshold]

    return [elements[i] for i in keep]


def _iou(a: Tuple, b: Tuple) -> float:
    ix1 = max(a[0], b[0]); iy1 = max(a[1], b[1])
    ix2 = min(a[2], b[2]); iy2 = min(a[3], b[3])
    inter = max(0, ix2 - ix1) * max(0, iy2 - iy1)
    if inter == 0:
        return 0.0
    ua = (a[2]-a[0])*(a[3]-a[1]) + (b[2]-b[0])*(b[3]-b[1]) - inter
    return inter / max(ua, 1)


# =============================================================================
# Main detector
# =============================================================================

class UIElementDetector:
    """
    Main facade — orchestrates OCR + contour detection, classification,
    annotation, crop export, and ROI datasheet export.
    Text hits are not grouped; every detection becomes an element.
    """

    # Manual padding (px) added around each detected text region
    BUTTON_PAD = 12
    LABEL_PAD  = 4

    # Known aircraft silhouette bounding box
    AIRCRAFT_ROI = (210, 95, 650, 700)

    def __init__(self):
        self._ocr_builder     = OCRElementBuilder()
        self._contour_builder = ContourElementBuilder()

    # -------------------------------------------------------------------------
    # Public API
    # -------------------------------------------------------------------------

    def detect(
        self,
        image_path:    str,
        out_dir:       str   = "./output",
        save_crops:    bool  = True,
        save_annotated: bool = True,
        save_xlsx:     bool  = True,
        save_csv:      bool  = True,
    ) -> List[UIElement]:
        """
        Full detection pipeline.

        Returns list of UIElement objects.
        Outputs:
            <out_dir>/annotated_<image_name>     — annotated element boundaries
            <out_dir>/text_annotated_<image_name>— detected text overlay
            <out_dir>/crops/<type>/<n>_<label>.png  — individual crops
            <out_dir>/roi_datasheet.xlsx         — Excel ROI table
            <out_dir>/roi_datasheet.csv          — CSV ROI table
            <out_dir>/text_datasheet.xlsx        — Excel text detections
            <out_dir>/text_datasheet.csv         — CSV text detections
        """
        img = cv2.imread(image_path)
        if img is None:
            raise FileNotFoundError(f"Cannot read: {image_path}")
        h, w = img.shape[:2]

        out_path = Path(out_dir)
        out_path.mkdir(parents=True, exist_ok=True)

        print(f"\n[1/5] Running OCR on {Path(image_path).name}  ({w}×{h}px) …")
        raw_ocr = self._ocr_builder.extract(img)
        # NOTE: NO merging – every raw OCR hit remains individual
        print(f"      {len(raw_ocr)} raw OCR hits (no grouping)")

        print("[2/5] Running contour detection …")
        contours = self._contour_builder.extract(img)
        print(f"      {len(contours)} contour candidates")

        print("[3/5] Classifying elements …")
        clf      = ElementClassifier(img, h, w)
        elements = self._build_elements(raw_ocr, contours, clf, img, w, h)
        print(f"      {len(elements)} elements after NMS")

        if save_crops:
            print("[4/5] Saving individual crops …")
            self._save_crops(img, elements, out_path)

        if save_annotated:
            print("[5a] Saving annotated image (boundaries) …")
            ann_path = out_path / f"annotated_{Path(image_path).name}"
            self._annotate(img.copy(), elements, str(ann_path))

            print("[5b] Saving annotated image (text) …")
            txt_ann_path = out_path / f"text_annotated_{Path(image_path).name}"
            self._annotate_text(img.copy(), raw_ocr, str(txt_ann_path))

        # Save main element datasheets
        if save_xlsx and _XLSX_AVAILABLE:
            print("[5c] Saving Excel ROI datasheet …")
            self._save_xlsx(elements, out_path / "roi_datasheet.xlsx",
                            Path(image_path).name)

        if save_csv:
            print("[5d] Saving CSV ROI datasheet …")
            self._save_csv(elements, out_path / "roi_datasheet.csv")

        # Save separate text datasheet (always, if OCR found anything)
        if raw_ocr:
            print("[5e] Saving text datasheet …")
            self._save_text_datasheet(raw_ocr, out_path)

        # Summary
        from collections import Counter
        counts = Counter(e.element_type for e in elements)
        print("\n  ── Detection summary ──────────────────────")
        for t, c in sorted(counts.items()):
            print(f"     {t:<22} {c:>3} element(s)")
        print(f"  ── Total: {len(elements)} ─────────────────────────\n")

        return elements

    # -------------------------------------------------------------------------
    # Build elements (no text grouping)
    # -------------------------------------------------------------------------

    def _build_elements(
        self,
        ocr_hits:  List[Dict],   # now raw OCR hits, not merged
        contours:  List[Dict],
        clf:       ElementClassifier,
        img:       np.ndarray,
        img_w:     int,
        img_h:     int,
    ) -> List[UIElement]:

        raw_elements: List[UIElement] = []
        idx = 0

        # ── From OCR hits ────────────────────────────────────────────────
        for hit in ocr_hits:
            x1, y1, x2, y2 = hit["x1"], hit["y1"], hit["x2"], hit["y2"]
            text = hit["text"].strip()

            # Expand tight OCR boxes to include button border
            pad = self.LABEL_PAD if len(text) > 20 else self.BUTTON_PAD
            x1e = max(0,     x1 - pad)
            y1e = max(0,     y1 - pad)
            x2e = min(img_w, x2 + pad)
            y2e = min(img_h, y2 + pad)

            etype = clf.classify(x1e, y1e, x2e, y2e, text)
            raw_elements.append(UIElement(
                idx          = idx,
                element_type = etype,
                label        = self._clean_label(text),
                x1=x1e, y1=y1e, x2=x2e, y2=y2e,
                confidence   = hit["conf"],
            ))
            idx += 1

        # ── Aircraft symbol (fixed ROI) ──────────────────────────────────
        ax1, ay1, ax2, ay2 = self.AIRCRAFT_ROI
        raw_elements.append(UIElement(
            idx          = idx,
            element_type = "AIRCRAFT_SYMBOL",
            label        = "Aircraft Silhouette",
            x1=ax1, y1=ay1, x2=ax2, y2=ay2,
            confidence   = 0.95,
            notes        = "Fixed ROI — graphical element",
        ))
        idx += 1

        # ── From contour hits (only add genuinely new regions) ───────────
        for c in contours:
            x1, y1, x2, y2 = c["x1"], c["y1"], c["x2"], c["y2"]
            # Skip if already covered by an OCR-derived element
            if any(_iou((x1,y1,x2,y2), (e.x1,e.y1,e.x2,e.y2)) > 0.35
                   for e in raw_elements):
                continue
            etype = clf.classify(x1, y1, x2, y2, "")
            if etype == "UNKNOWN":
                continue
            raw_elements.append(UIElement(
                idx          = idx,
                element_type = etype,
                label        = f"{etype.lower()}_region",
                x1=x1, y1=y1, x2=x2, y2=y2,
                confidence   = c["conf"],
                notes        = "contour-detected",
            ))
            idx += 1

        # ── NMS ─────────────────────────────────────────────────────────
        final = nms(raw_elements, iou_threshold=0.50)

        # Re-index
        for i, e in enumerate(final):
            e.idx = i + 1

        return final

    # -------------------------------------------------------------------------
    # Annotated image (boundaries)
    # -------------------------------------------------------------------------

    def _annotate(self, img: np.ndarray,
                  elements: List[UIElement], out_path: str) -> None:
        overlay = img.copy()

        for e in elements:
            color = TYPE_COLORS.get(e.element_type, (100, 100, 100))
            # Semi-transparent fill
            cv2.rectangle(overlay, (e.x1, e.y1), (e.x2, e.y2), color, -1)

        # Blend overlay
        cv2.addWeighted(overlay, 0.18, img, 0.82, 0, img)

        for e in elements:
            color = TYPE_COLORS.get(e.element_type, (100, 100, 100))
            # Bounding box
            cv2.rectangle(img, (e.x1, e.y1), (e.x2, e.y2), color, 2)

            # Index number tag
            tag_x, tag_y = e.x1, max(12, e.y1 - 4)
            tag_text     = str(e.idx)
            (tw, th), _  = cv2.getTextSize(
                tag_text, cv2.FONT_HERSHEY_SIMPLEX, 0.42, 1)
            cv2.rectangle(img,
                          (tag_x, tag_y - th - 3),
                          (tag_x + tw + 4, tag_y + 2),
                          color, -1)
            cv2.putText(img, tag_text,
                        (tag_x + 2, tag_y - 1),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.42,
                        (0, 0, 0), 1, cv2.LINE_AA)

        cv2.imwrite(out_path, img)
        print(f"      Saved → {out_path}")

    # -------------------------------------------------------------------------
    # Text-annotated image
    # -------------------------------------------------------------------------

    def _annotate_text(self, img: np.ndarray,
                       ocr_hits: List[Dict], out_path: str) -> None:
        overlay = img.copy()

        for hit in ocr_hits:
            x1, y1, x2, y2 = hit["x1"], hit["y1"], hit["x2"], hit["y2"]
            # Light cyan background for text regions
            cv2.rectangle(overlay, (x1, y1), (x2, y2), (255, 255, 0), -1)

        cv2.addWeighted(overlay, 0.15, img, 0.85, 0, img)

        for hit in ocr_hits:
            x1, y1, x2, y2 = hit["x1"], hit["y1"], hit["x2"], hit["y2"]
            text = hit["text"]
            font_scale = 0.45
            cv2.putText(img, text, (x1, max(y1-4, 10)),
                        cv2.FONT_HERSHEY_SIMPLEX, font_scale,
                        (0, 0, 0), 2, cv2.LINE_AA)

        cv2.imwrite(out_path, img)
        print(f"      Saved → {out_path}")

    # -------------------------------------------------------------------------
    # Crop export
    # -------------------------------------------------------------------------

    def _save_crops(self, img: np.ndarray,
                    elements: List[UIElement],
                    out_path: Path) -> None:
        crops_dir = out_path / "crops"
        crops_dir.mkdir(exist_ok=True)

        for e in elements:
            type_dir = crops_dir / e.element_type
            type_dir.mkdir(exist_ok=True)

            x1, y1 = max(0, e.x1), max(0, e.y1)
            x2, y2 = min(img.shape[1], e.x2), min(img.shape[0], e.y2)
            if x2 <= x1 or y2 <= y1:
                continue

            crop = img[y1:y2, x1:x2]
            safe = re.sub(r"[^\w]", "_", e.label)[:40]
            fname = f"{e.idx:03d}_{safe}.png"
            fpath = type_dir / fname
            cv2.imwrite(str(fpath), crop)
            e.crop_path = str(fpath)

        total = sum(1 for e in elements if e.crop_path)
        print(f"      {total} crops saved under {crops_dir}/")

    # -------------------------------------------------------------------------
    # Text datasheet (separate from element datasheet)
    # -------------------------------------------------------------------------

    def _save_text_datasheet(self, ocr_hits: List[Dict],
                             out_path: Path) -> None:
        """Save a CSV and XLSX of raw OCR text detections."""
        # CSV
        csv_path = out_path / "text_datasheet.csv"
        headers = ["index", "text", "x1", "y1", "x2", "y2", "confidence"]
        with open(str(csv_path), "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for i, hit in enumerate(ocr_hits, 1):
                writer.writerow({
                    "index": i,
                    "text": hit["text"],
                    "x1": hit["x1"], "y1": hit["y1"],
                    "x2": hit["x2"], "y2": hit["y2"],
                    "confidence": round(hit["conf"], 3),
                })
        print(f"      Saved → {csv_path}")

        if _XLSX_AVAILABLE:
            xlsx_path = out_path / "text_datasheet.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "OCR Text Detections"
            ws.append(headers)
            for i, hit in enumerate(ocr_hits, 1):
                ws.append([i, hit["text"], hit["x1"], hit["y1"],
                           hit["x2"], hit["y2"], round(hit["conf"], 3)])
            # Simple formatting
            header_fill = PatternFill("solid", fgColor="1A2B4A")
            header_font = Font(bold=True, color="FFFFFF")
            for col in range(1, 8):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            wb.save(str(xlsx_path))
            print(f"      Saved → {xlsx_path}")

    # -------------------------------------------------------------------------
    # Excel export (element datasheet)
    # -------------------------------------------------------------------------

    def _save_xlsx(self, elements: List[UIElement],
                   out_path: Path, source_image: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ROI Datasheet"

        # ── Header ──────────────────────────────────────────────────────
        headers = [
            "Index", "Type", "Label",
            "X1 (px)", "Y1 (px)", "X2 (px)", "Y2 (px)",
            "Width (px)", "Height (px)", "Area (px²)",
            "Confidence", "Notes", "Crop Path",
        ]
        header_fill  = PatternFill("solid", fgColor="1A2B4A")
        header_font  = Font(bold=True, color="FFFFFF", size=10)
        header_align = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        thin = Side(style="thin", color="444444")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill  = header_fill
            cell.font  = header_font
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 28

        # Type-based row fill colors
        type_fills = {
            "BUTTON_ACTIVE":    "C8F5D5",
            "BUTTON_NORMAL":    "D6E8FF",
            "BUTTON_ALERT":     "FFE0B2",
            "TAB_ACTIVE":       "B9F6CA",
            "TAB_INACTIVE":     "EEEEEE",
            "SECTION_LABEL":    "FFF9C4",
            "TITLE":            "F3E5F5",
            "DISPLAY_FIELD":    "EDE7F6",
            "AIRCRAFT_SYMBOL":  "E0F7FA",
            "NAVBAR_ITEM":      "FCE4EC",
            "PANEL_HEADER":     "E1F5FE",
            "TEXT_LABEL":       "F5F5F5",
            "ATTENDANT_ENTRY":  "FFF3E0",
            "SLIDER_WIDGET":    "E8F5E9",
            "UNKNOWN":          "FAFAFA",
        }

        # ── Data rows ───────────────────────────────────────────────────
        for row_idx, e in enumerate(elements, 2):
            fill_hex = type_fills.get(e.element_type, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_hex)
            row_data = [
                e.idx,
                e.element_type,
                e.label,
                e.x1, e.y1, e.x2, e.y2,
                e.width, e.height, e.area,
                round(e.confidence, 3),
                e.notes,
                e.crop_path,
            ]
            ws.append(row_data)
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill   = row_fill
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if col not in (3, 13) else "left",
                    vertical="center",
                )
                if col == 11:                        # confidence
                    cell.number_format = "0.000"
                if col == 10:                        # area
                    cell.number_format = "#,##0"

        # ── Column widths ────────────────────────────────────────────────
        col_widths = [7, 20, 30, 9, 9, 9, 9, 10, 11, 12, 12, 25, 50]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # ── Legend sheet ────────────────────────────────────────────────
        ls = wb.create_sheet("Legend")
        ls.append(["Type", "Description", "Color code"])
        ls.column_dimensions["A"].width = 22
        ls.column_dimensions["B"].width = 45
        ls.column_dimensions["C"].width = 14

        descriptions = {
            "BUTTON_ACTIVE":    "Pressable button — currently active/selected (green fill)",
            "BUTTON_NORMAL":    "Pressable button — default/inactive state",
            "BUTTON_ALERT":     "Alert or warning button (orange fill)",
            "TAB_ACTIVE":       "Navigation tab — currently selected",
            "TAB_INACTIVE":     "Navigation tab — not selected",
            "SECTION_LABEL":    "Non-interactive section heading text",
            "TITLE":            "Page or screen title",
            "DISPLAY_FIELD":    "Read-only value display field",
            "AIRCRAFT_SYMBOL":  "Graphical aircraft silhouette (non-interactive)",
            "NAVBAR_ITEM":      "Bottom navigation bar item",
            "PANEL_HEADER":     "Header row inside a data panel",
            "TEXT_LABEL":       "Small informational text label",
            "ATTENDANT_ENTRY":  "Attendant zone row inside the entries panel",
            "SLIDER_WIDGET":    "Brightness / dimming slider control",
            "UNKNOWN":          "Unclassified element",
        }
        for t, desc in descriptions.items():
            ls.append([t, desc, type_fills.get(t, "FFFFFF")])
            row = ls.max_row
            ls.cell(row=row, column=1).fill = PatternFill(
                "solid", fgColor=type_fills.get(t, "FFFFFF"))

        # ── Meta sheet ──────────────────────────────────────────────────
        ms = wb.create_sheet("Meta")
        ms.append(["Source image",    source_image])
        ms.append(["Total elements",  len(elements)])
        ms.append(["Generated by",    "UIElementDetector — Cyient Ltd."])
        ms.append(["Script",          "ui_element_detector.py"])
        ms.column_dimensions["A"].width = 18
        ms.column_dimensions["B"].width = 50

        wb.save(str(out_path))
        print(f"      Saved → {out_path}")

    # -------------------------------------------------------------------------
    # CSV export (element datasheet)
    # -------------------------------------------------------------------------

    def _save_csv(self, elements: List[UIElement], out_path: Path) -> None:
        headers = ["index", "type", "label",
                   "x1", "y1", "x2", "y2",
                   "width", "height", "area",
                   "confidence", "notes", "crop_path"]
        with open(str(out_path), "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for e in elements:
                w.writerow({
                    "index":      e.idx,
                    "type":       e.element_type,
                    "label":      e.label,
                    "x1":         e.x1,  "y1": e.y1,
                    "x2":         e.x2,  "y2": e.y2,
                    "width":      e.width,
                    "height":     e.height,
                    "area":       e.area,
                    "confidence": round(e.confidence, 3),
                    "notes":      e.notes,
                    "crop_path":  e.crop_path,
                })
        print(f"      Saved → {out_path}")

    # -------------------------------------------------------------------------
    # Helpers
    # -------------------------------------------------------------------------

    @staticmethod
    def _clean_label(text: str) -> str:
        return re.sub(r"\s+", " ", text.strip())[:60]


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Detect and annotate every UI element in an HSI screenshot."
    )
    ap.add_argument("image",     help="Path to the reference HSI screenshot")
    ap.add_argument("--out-dir", default="./output",
                    help="Output directory (default: ./output)")
    ap.add_argument("--no-crops",     action="store_true")
    ap.add_argument("--no-annotated", action="store_true")
    ap.add_argument("--no-xlsx",      action="store_true")
    ap.add_argument("--no-csv",       action="store_true")
    args = ap.parse_args()

    det = UIElementDetector()
    det.detect(
        image_path    = args.image,
        out_dir       = args.out_dir,
        save_crops    = not args.no_crops,
        save_annotated= not args.no_annotated,
        save_xlsx     = not args.no_xlsx,
        save_csv      = not args.no_csv,
    )